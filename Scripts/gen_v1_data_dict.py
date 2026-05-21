"""Generate FDD v1 xlsx and yaml data dictionaries (empty values, variable names only)"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import yaml

PROJECT = Path(__file__).resolve().parent.parent

# ===== XLSX =====
wb = openpyxl.Workbook()
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")

def make_kv_sheet(wb, name, fields, is_first=False):
    if is_first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 55
    ws.append(["字段编码", "值（留空待填）"])
    ws['A1'].font = header_font; ws['B1'].font = header_font
    ws['A1'].fill = header_fill; ws['B1'].fill = header_fill
    for field in fields:
        ws.append([field, ""])
    return ws

def make_table_sheet(wb, name, columns):
    ws = wb.create_sheet(name)
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22
    return ws

# date sheets
make_kv_sheet(wb, "date.全局", [
    "公司名称", "公司简称", "报告基准日", "报告基准日_大写",
    "报告文号", "项目名称",
], is_first=True)

make_kv_sheet(wb, "date.历史沿革", [
    "公司名称", "统一社会信用代码", "注册地址", "法定代表人",
    "公司类型", "注册资本", "经营范围", "成立日期", "营业期限",
])

make_kv_sheet(wb, "date.项目概况", [
    "上网模式", "不含税总价", "不含税总价单价",
    "地址", "竣工容量", "装机容量",
])

# form sheets
make_table_sheet(wb, "form.释义", ["全称", "简称"])
make_table_sheet(wb, "form.股权结构", ["股东名称", "认缴出资额", "认缴比例", "实缴出资额", "实缴比例"])
make_table_sheet(wb, "form.股东信息", ["股东名称", "股东介绍"])
make_table_sheet(wb, "form.公司设立", ["股东名称", "认缴出资额", "认缴比例", "实缴出资额", "实缴比例"])

xlsx_path = PROJECT / "data" / "FDDv1_数据字典.xlsx"
wb.save(str(xlsx_path))
print("XLSX saved:", xlsx_path)

# ===== YAML =====
yaml_data = {
    "date": {
        "全局": {
            "公司名称": "",
            "公司简称": "",
            "报告基准日": "",
            "报告基准日_大写": "",
            "报告文号": "",
            "项目名称": "",
        },
        "历史沿革": {
            "公司名称": "",
            "统一社会信用代码": "",
            "注册地址": "",
            "法定代表人": "",
            "公司类型": "",
            "注册资本": "",
            "经营范围": "",
            "成立日期": "",
            "营业期限": "",
        },
        "项目概况": {
            "上网模式": "",
            "不含税总价": "",
            "不含税总价单价": "",
            "地址": "",
            "竣工容量": "",
            "装机容量": "",
        },
    },
    "form": {
        "释义": [],
        "股权结构": [],
        "股东信息": [],
        "公司设立": [],
    }
}

yaml_path = PROJECT / "data" / "FDDv1_数据字典.yaml"
with open(str(yaml_path), 'w', encoding='utf-8') as f:
    f.write("# FDD v1 数据字典 — 变量名留空待填\n")
    f.write("# 模板: templates/FDD初稿 v1.docx\n\n")
    yaml.dump(yaml_data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
print("YAML saved:", yaml_path)
print("Done.")
